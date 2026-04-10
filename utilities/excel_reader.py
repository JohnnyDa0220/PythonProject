"""
Author Name: ROBIN MAHANTA
Module: excel_reader.py
Purpose: Provides utility functions to read test data from Excel files.
Description: This module contains the read_excel function which loads data from Excel workbooks
             and retrieves cell values for use in test parameterization.
"""

from openpyxl import load_workbook
import os

def read_excel(path, sheet_name, row, col):
    """
    Author Name: ROBIN MAHANTA
    Function: read_excel
    Purpose: Reads a cell value from an Excel file.
    Description: Loads an Excel workbook, accesses the specified sheet, 
                 and returns the value from the specified row and column.
    Parameters:
        - path (str): Full path to the Excel file (.xlsx)
        - sheet_name (str): Name of the sheet to read from
        - row (int): Row number (1-indexed)
        - col (int): Column number (1-indexed)
    Return Type: Cell value (str, int, float, etc.)
    Raises:
        - FileNotFoundError: If the Excel file doesn't exist
        - ValueError: If the file format is not .xlsx
    """
    if not os.path.exists(path):
        raise FileNotFoundError(f"Excel file not found: {path}")

    if not path.endswith(".xlsx"):
        raise ValueError(f"Invalid Excel file format (must be .xlsx): {path}")

    workbook = load_workbook(path)
    sheet = workbook[sheet_name]
    return sheet.cell(row=row, column=col).value
